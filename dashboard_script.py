"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  PROJECT PORTFOLIO DASHBOARD GENERATOR  v3.0                               ║
║  + Auto column detection  (no fixed column indices needed)                 ║
║  + Fuzzy duplicate client-name detection  (rapidfuzz)                      ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  INPUT  : Any Excel file with columns (name variations handled):           ║
║           Client Name | PO/WO Reference | Work Value | Billed Amount       ║
║  OUTPUT : 3-sheet Excel dashboard + duplicate-flag section in Analysis     ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  INSTALL: pip install pandas openpyxl rapidfuzz                            ║
║  RUN    : python generate_portfolio_dashboard.py                           ║
║           python generate_portfolio_dashboard.py myfile.xlsx out.xlsx     ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

import re
import sys
import logging
import argparse
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  ❶  CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Config:
    """
    All tuneable settings in one place.
    Column positions are now discovered automatically; override only if
    auto-detection fails on a non-standard file.
    """

    # ── File paths
    input_file:  str = "Raw_data.xlsx"
    output_file: str = "Project_Portfolio_Dashboard.xlsx"

    # ── Auto column detection
    # Canonical names for each required field.  The detector normalises both
    # the header cell and these strings before comparing, so minor typos /
    # casing differences are handled automatically.
    col_aliases: Dict[str, List[str]] = field(default_factory=lambda: {
        "client": [
            "client name", "client", "customer name", "customer",
            "party name", "company name",
        ],
        "po": [
            "po/wo no", "po/wo no.", "po no", "wo no", "po number",
            "work order", "purchase order", "project reference",
            "po reference", "po / wo reference", "order no", "order number",
            "po-wo reference", "po/wo reference",
        ],
        "wv": [
            "work order value", "total work value", "work value",
            "contract value", "po value", "order value",
            "total value", "project value", "amount",
        ],
        "billed": [
            "amt executed till date", "billed amount", "executed amount",
            "amount executed", "invoiced amount", "billed", "executed",
            "amount billed", "amount invoiced", "billing amount",
        ],
    })

    # Manual column overrides (0-indexed).  Set a key to an int to skip
    # auto-detection for that field.  Leave None to use auto-detection.
    col_override: Dict[str, Optional[int]] = field(default_factory=lambda: {
        "client": None,
        "po":     None,
        "wv":     None,
        "billed": None,
    })

    # ── Header-row detection
    # Max rows to scan from the top of the file looking for the header row.
    header_scan_limit: int = 10

    # ── Client-name normalisation  raw_name -> clean_name
    client_name_map: Dict[str, str] = field(default_factory=lambda: {
        # Add pairs as needed.  Leave empty ({}) if data is already clean.
        # "Old messy name": "Clean canonical name",
    })

    # ── Fuzzy duplicate detection
    # Pairs with token_sort_ratio >= this threshold are flagged as potential
    # duplicates.  Range 0–100; 85 is a good balance for company names.
    fuzzy_threshold: int = 85

    # ── Status bands  (lo_exclusive, hi_inclusive, label)
    status_bands: List[Tuple[float, float, str]] = field(default_factory=lambda: [
        (0.00,  0.00,  "Not Started"),
        (0.00,  0.25,  "Early Stage"),
        (0.25,  0.75,  "Active Execution"),
        (0.75,  1.00,  "Near Completion"),
        (1.00, 999.0,  "Over Execution"),
    ])


DEFAULT_CONFIG = Config()


# ══════════════════════════════════════════════════════════════════════════════
#  ❷  STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

ALT_ROW_BG = "EBF5FB"

CAT_COLORS: Dict[str, Tuple[str, str]] = {
    "Strategic":          ("C6EFCE", "375623"),
    "Attention Required": ("FFEB9C", "9C5700"),
    "Stable Clients":     ("DDEBF7", "1F4E79"),
    "Low Priority":       ("FCE4D6", "843C0C"),
}

CAT_DESC: Dict[str, str] = {
    "Strategic":          "High Revenue + High Execution — Core performers, maintain focus",
    "Attention Required": "High Revenue + Low Execution — High-value clients needing intervention",
    "Stable Clients":     "Low Revenue + High Execution — Reliable, efficient smaller projects",
    "Low Priority":       "Low Revenue + Low Execution — Review, escalate or defer",
}

CATS: List[str] = ["Strategic", "Attention Required", "Stable Clients", "Low Priority"]

DUPE_BG  = "FFF2CC"   # light amber — duplicate warning rows
DUPE_HDR = "ED7D31"   # orange — duplicate section header
WARN_RED = "C00000"   # red text for similarity score

_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════════════════════════════════════
#  ❸  DOMAIN TYPES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ProjectRecord:
    """One cleaned row from the source file."""
    client: str
    po:     str
    wv:     float
    ba:     float


@dataclass
class DuplicatePair:
    """Two client names that exceed the fuzzy-match threshold."""
    name_a:     str
    name_b:     str
    similarity: float   # 0–100


@dataclass
class ColumnMap:
    """Resolved 0-indexed column positions after auto-detection."""
    client: int
    po:     int
    wv:     int
    billed: int
    header_row: int     # 0-indexed row that contained the headers


# ══════════════════════════════════════════════════════════════════════════════
#  ❹  STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _font(bold=False, size=10, color="000000", italic=False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def col_w(ws: Worksheet, col_index: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col_index)].width = width

def apply_cell(ws, row, col, value=None, *, bold=False, size=10,
               color="000000", bg=None, align_h="left", align_v="center",
               fmt=None, wrap=False, border=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = _align(h=align_h, v=align_v, wrap=wrap)
    if border:
        c.border = _THIN_BORDER
    if bg:
        c.fill = _fill(bg)
    if fmt:
        c.number_format = fmt
    return c

def title_banner(ws, row, span, text, bg="1F4E79"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    apply_cell(ws, row, 1, text, bold=True, size=13, color="FFFFFF",
               bg=bg, align_h="center", border=False)
    ws.row_dimensions[row].height = 32

def section_title(ws, row, span, text, bg="D6E4F0", fg="1F4E79"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    apply_cell(ws, row, 1, text, bold=True, size=11, color=fg,
               bg=bg, align_h="left", border=False)
    ws.row_dimensions[row].height = 22

def header_row(ws, row, labels, bgs, height=36):
    for col, (lbl, bg) in enumerate(zip(labels, bgs), 1):
        apply_cell(ws, row, col, lbl, bold=True, size=10,
                   color="FFFFFF", bg=bg, align_h="center", wrap=True)
    ws.row_dimensions[row].height = height

def dark_total_row(ws, row, n_cols):
    for co in range(1, n_cols + 1):
        apply_cell(ws, row, co, bold=True, color="FFFFFF",
                   bg="1F4E79", align_h="center")


# ══════════════════════════════════════════════════════════════════════════════
#  ❺  FORMULA BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def status_excel_formula(pct_cell: str) -> str:
    p = pct_cell
    return (f'=IF({p}=0,"Not Started",'
            f'IF({p}<=0.25,"Early Stage",'
            f'IF({p}<=0.75,"Active Execution",'
            f'IF({p}<1,"Near Completion","Over Execution"))))')

def category_excel_formula(rev_cell, exe_cell, rev_thr, exe_thr) -> str:
    r, e, rt, et = rev_cell, exe_cell, rev_thr, exe_thr
    return (f'=IF(AND({r}>={rt},{e}>={et}),"Strategic",'
            f'IF(AND({r}>={rt},{e}<{et}),"Attention Required",'
            f'IF(AND({r}<{rt},{e}>={et}),"Stable Clients","Low Priority")))')


# ══════════════════════════════════════════════════════════════════════════════
#  ❻  AUTO COLUMN DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def _normalise_header(text) -> str:
    """Lowercase, collapse whitespace, strip punctuation for robust matching."""
    if pd.isna(text):
        return ""
    s = str(text).lower().strip()
    s = re.sub(r"[/\-_\.\(\)]", " ", s)   # punctuation → space
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _match_column(header_norm: str, aliases: List[str]) -> bool:
    """
    Return True if `header_norm` matches any alias.
    Uses exact match first, then substring match.
    """
    for alias in aliases:
        alias_norm = _normalise_header(alias)
        if header_norm == alias_norm:
            return True
        if alias_norm in header_norm or header_norm in alias_norm:
            return True
    return False


def detect_columns(df: pd.DataFrame, cfg: Config) -> ColumnMap:
    """
    Scan rows 0..header_scan_limit to find the header row, then match
    each column to the four required fields using alias lists.

    Returns a ColumnMap with resolved 0-indexed positions.
    Raises ValueError if any required column cannot be found.
    """
    header_row_idx = None
    header_series  = None

    # ── Step 1: find the header row ──────────────────────────────────────────
    for row_idx in range(min(cfg.header_scan_limit, len(df))):
        row_vals = df.iloc[row_idx]
        normed   = [_normalise_header(v) for v in row_vals]
        # A row qualifies as the header if it matches at least 2 required fields
        matches = 0
        for field_name, aliases in cfg.col_aliases.items():
            if cfg.col_override.get(field_name) is not None:
                matches += 1   # manually set — counts as found
                continue
            for val in normed:
                if val and _match_column(val, aliases):
                    matches += 1
                    break
        if matches >= 2:
            header_row_idx = row_idx
            header_series  = normed
            break

    if header_row_idx is None:
        raise ValueError(
            f"Could not find a header row in the first {cfg.header_scan_limit} rows. "
            "Check that the file contains the expected column names."
        )

    log.info("Header row detected at 0-indexed row %d.", header_row_idx)

    # ── Step 2: match each field ──────────────────────────────────────────────
    resolved: Dict[str, int] = {}
    unmatched: List[str]     = []

    for field_name, aliases in cfg.col_aliases.items():
        override = cfg.col_override.get(field_name)
        if override is not None:
            resolved[field_name] = override
            log.info("  %-8s → col %d  (manual override)", field_name, override)
            continue

        found = None
        for col_idx, val in enumerate(header_series):
            if val and _match_column(val, aliases):
                found = col_idx
                break

        if found is None:
            unmatched.append(field_name)
        else:
            raw_header = str(df.iloc[header_row_idx, found]).strip()
            log.info("  %-8s → col %d  ('%s')", field_name, found, raw_header)
            resolved[field_name] = found

    if unmatched:
        # Build a helpful error with what was actually found vs expected
        found_headers = [
            str(df.iloc[header_row_idx, c]).strip()
            for c in range(df.shape[1])
            if not pd.isna(df.iloc[header_row_idx, c])
        ]
        raise ValueError(
            f"Could not auto-detect columns for: {unmatched}.\n"
            f"Headers found in row {header_row_idx}: {found_headers}\n"
            f"Add the actual header text to the matching alias list in Config.col_aliases, "
            "or use col_override to set the column index manually."
        )

    return ColumnMap(
        client     = resolved["client"],
        po         = resolved["po"],
        wv         = resolved["wv"],
        billed     = resolved["billed"],
        header_row = header_row_idx,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  ❼  DATA LAYER  —  load · clean · validate · categorise
# ══════════════════════════════════════════════════════════════════════════════

def _clean_str(s) -> str:
    if pd.isna(s):
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\n", " ")).strip()


def _safe_float(val, row_num: int, field_name: str) -> float:
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        raise ValueError(
            f"Row {row_num}: expected number in '{field_name}', got '{val}'."
        )


def load_records(filepath: str | Path,
                 cfg: Config) -> Tuple[List[ProjectRecord], ColumnMap]:
    """
    Read source Excel → auto-detect columns → clean rows → return records.

    Returns
    -------
    (records, col_map)
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path.resolve()}")
    if path.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
        raise ValueError(f"Expected .xlsx/.xls/.xlsm, got '{path.name}'")

    log.info("Reading  : %s", path.name)
    try:
        df = pd.read_excel(path, header=None)
    except Exception as exc:
        raise ValueError(f"Cannot read '{path.name}': {exc}") from exc

    if df.empty:
        raise ValueError("Source file is empty.")

    # ── Auto-detect columns
    col_map = detect_columns(df, cfg)

    # ── Parse data rows
    data_block = df.iloc[col_map.header_row + 1:].reset_index(drop=True)
    records: List[ProjectRecord] = []
    skipped = 0

    for raw_idx, row in data_block.iterrows():
        excel_row = raw_idx + col_map.header_row + 2

        client = _clean_str(row.iloc[col_map.client])
        if not client or client.lower() in ("client name", "client"):
            skipped += 1
            continue

        # Apply name normalisation map
        client = cfg.client_name_map.get(client, client).replace("\n", " ").strip()
        po     = _clean_str(row.iloc[col_map.po])

        try:
            wv     = _safe_float(row.iloc[col_map.wv],     excel_row, "Work Value")
            billed = _safe_float(row.iloc[col_map.billed], excel_row, "Billed Amount")
        except ValueError as exc:
            log.warning("%s — row skipped.", exc)
            skipped += 1
            continue

        records.append(ProjectRecord(client=client, po=po, wv=wv, ba=billed))

    if skipped:
        log.info("Skipped  : %d row(s) (blank / header / bad data).", skipped)
    if not records:
        raise ValueError(
            "No valid data rows found. Check Config.col_aliases or col_override."
        )

    return records, col_map


def unique_clients(records: List[ProjectRecord]) -> List[str]:
    seen: set = set()
    result: List[str] = []
    for rec in records:
        if rec.client not in seen:
            seen.add(rec.client)
            result.append(rec.client)
    return result


def compute_categories(records: List[ProjectRecord]) -> Dict[str, List[int]]:
    """Assign each record to a quadrant category (mirrors Analysis formula logic)."""
    n        = len(records)
    total_wv = sum(r.wv for r in records) or 1.0
    rev_thr  = 1.0 / n
    exe_thr  = sum(r.ba / r.wv if r.wv > 0 else 0.0 for r in records) / n

    buckets: Dict[str, List[int]] = {c: [] for c in CATS}
    for idx, rec in enumerate(records, 1):
        rev  = rec.wv / total_wv
        comp = rec.ba / rec.wv if rec.wv > 0 else 0.0
        hi_r, hi_e = rev >= rev_thr, comp >= exe_thr
        if   hi_r and hi_e:     cat = "Strategic"
        elif hi_r and not hi_e: cat = "Attention Required"
        elif not hi_r and hi_e: cat = "Stable Clients"
        else:                   cat = "Low Priority"
        buckets[cat].append(idx)

    return buckets


# ══════════════════════════════════════════════════════════════════════════════
#  ❽  FUZZY DUPLICATE DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def detect_duplicate_clients(clients: List[str],
                              threshold: int) -> List[DuplicatePair]:
    """
    Compare every unique pair of client names using rapidfuzz.token_sort_ratio.

    token_sort_ratio is chosen because it:
    - Is insensitive to word order  ("Tata Power Ltd" vs "Ltd Tata Power")
    - Handles common prefix/suffix differences well  ("THE TATA..." vs "TATA...")
    - Normalises case automatically

    Returns pairs with similarity >= threshold, sorted descending by score.
    """
    pairs: List[DuplicatePair] = []
    n = len(clients)
    for i in range(n):
        for j in range(i + 1, n):
            score = fuzz.token_sort_ratio(clients[i], clients[j])
            if score >= threshold:
                pairs.append(DuplicatePair(
                    name_a=clients[i],
                    name_b=clients[j],
                    similarity=score,
                ))
    pairs.sort(key=lambda p: p.similarity, reverse=True)

    if pairs:
        log.warning(
            "⚠  Found %d potential duplicate client name pair(s) "
            "(threshold: %d%%):", len(pairs), threshold
        )
        for p in pairs:
            log.warning("   %.1f%%  '%s'  ↔  '%s'", p.similarity, p.name_a, p.name_b)
    else:
        log.info("✓  No duplicate client names detected (threshold: %d%%).", threshold)

    return pairs


# ══════════════════════════════════════════════════════════════════════════════
#  ❾  SHEET 1  —  RAW DATA  (unchanged logic)
# ══════════════════════════════════════════════════════════════════════════════

def _raw_data_rows(ws, records):
    for idx, rec in enumerate(records, 1):
        r  = idx + 2
        bg = ALT_ROW_BG if idx % 2 == 0 else "FFFFFF"
        apply_cell(ws, r, 1, idx,        bg=bg, align_h="center")
        apply_cell(ws, r, 2, rec.client, bg=bg)
        apply_cell(ws, r, 3, rec.po,     bg=bg)
        apply_cell(ws, r, 4, rec.wv,     bg=bg, align_h="right", fmt="#,##0.00")
        apply_cell(ws, r, 5, rec.ba,     bg=bg, align_h="right", fmt="#,##0.00")
        c6 = ws.cell(r, 6, f"=IFERROR(E{r}/D{r},0)")
        c6.number_format = "0.0%"; c6.font = _font()
        c6.alignment = _align(h="center"); c6.border = _THIN_BORDER; c6.fill = _fill(bg)
        c7 = ws.cell(r, 7, status_excel_formula(f"F{r}"))
        c7.font = _font(); c7.alignment = _align(h="center")
        c7.border = _THIN_BORDER; c7.fill = _fill(bg)


def _raw_totals_row(ws, last_row):
    tr = last_row + 1
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=3)
    dark_total_row(ws, tr, 7)
    ws.cell(tr, 1).value     = "TOTAL"
    ws.cell(tr, 1).alignment = _align(h="right")
    for col, rng in [(4, f"D3:D{last_row}"), (5, f"E3:E{last_row}")]:
        ws.cell(tr, col).value         = f"=SUM({rng})"
        ws.cell(tr, col).number_format = "#,##0.00"
        ws.cell(tr, col).alignment     = _align(h="right")


def build_raw_data_sheet(ws: Worksheet,
                          records: List[ProjectRecord]) -> int:
    ws.title        = "Raw Data"
    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "1F4E79"
    last_row = len(records) + 2

    title_banner(ws, 1, 7, "RAW DATA  —  Project Portfolio")
    header_row(ws, 2,
               ["Sr. No.", "Client Name", "Project / PO-WO Reference",
                "Total Work Value (₹)", "Billed Amount (₹)",
                "Completion %", "Status"],
               ["1F4E79"] * 7)
    _raw_data_rows(ws, records)
    _raw_totals_row(ws, last_row)

    for col, width in [(1,8),(2,38),(3,50),(4,22),(5,22),(6,14),(7,18)]:
        col_w(ws, col, width)

    return last_row


# ══════════════════════════════════════════════════════════════════════════════
#  ❿  SHEET 2  —  ANALYSIS  (existing tables + NEW duplicate section)
# ══════════════════════════════════════════════════════════════════════════════

def _threshold_params(ws, last_raw) -> Tuple[str, str]:
    section_title(ws, 3, 9,
                  "  📌  THRESHOLD PARAMETERS  (used for category assignment)")
    params = [
        ("Revenue High Threshold  (Avg Revenue Share per Project)",
         f"=1/COUNTA('Raw Data'!B3:B{last_raw})"),
        ("Execution High Threshold  (Avg Completion % across all projects)",
         f"=AVERAGE('Raw Data'!F3:F{last_raw})"),
    ]
    for pi, (lbl, fml) in enumerate(params):
        pr = 4 + pi
        ws.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=4)
        for co in range(1, 10):
            apply_cell(ws, pr, co, bg="D6E4F0", color="1F4E79", bold=True)
        ws.cell(pr, 1).value     = lbl
        ws.cell(pr, 1).alignment = _align(h="right")
        cv = ws.cell(pr, 5, fml)
        cv.font = _font(bold=True, color="375623"); cv.number_format = "0.00%"
        cv.alignment = _align(h="center"); cv.border = _THIN_BORDER
        cv.fill = _fill("C6EFCE")
    return "$E$4", "$E$5"


def _table1_status_distribution(ws, clients, last_raw, start_row) -> int:
    section_title(ws, start_row, 9,
                  "  TABLE 1 :  CLIENT-WISE PROJECT STATUS DISTRIBUTION")
    note_r = start_row + 1
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=9)
    nc = ws.cell(note_r, 1,
        "Status Definitions:  Not Started = 0%  |  Early Stage = 0–25%  |  "
        "Active Execution = 25–75%  |  Near Completion = 75–99%  |  Over Execution ≥ 100%")
    nc.font = Font(name="Arial", italic=True, size=9, color="595959")
    nc.fill = _fill("F2F2F2"); nc.alignment = _align(h="left")
    ws.row_dimensions[note_r].height = 18

    T1_HDR = start_row + 2
    header_row(ws, T1_HDR,
               ["Client Name","Not Started","Early Stage","Active Execution",
                "Near Completion","Over Execution","Total Projects"],
               ["1F4E79","FF0000","ED7D31","4472C4","70AD47","A61C00","1F4E79"])

    T1_DATA = T1_HDR + 1
    STATUS_LABELS = ["Not Started","Early Stage","Active Execution",
                     "Near Completion","Over Execution"]
    for ci, client in enumerate(clients):
        r  = T1_DATA + ci
        bg = ALT_ROW_BG if ci % 2 == 0 else "FFFFFF"
        apply_cell(ws, r, 1, client, bg=bg)
        for si, slbl in enumerate(STATUS_LABELS, 2):
            f = (f"=COUNTIFS('Raw Data'!$B$3:'Raw Data'!$B${last_raw},A{r},"
                 f"'Raw Data'!$G$3:'Raw Data'!$G${last_raw},\"{slbl}\")")
            c = ws.cell(r, si, f)
            c.font = _font(); c.alignment = _align(h="center")
            c.border = _THIN_BORDER; c.fill = _fill(bg)
        tc = ws.cell(r, 7, f"=SUM(B{r}:F{r})")
        tc.font = _font(bold=True); tc.alignment = _align(h="center")
        tc.border = _THIN_BORDER; tc.fill = _fill(bg)

    T1_TOT = T1_DATA + len(clients)
    for co in range(1, 8):
        cl  = get_column_letter(co)
        val = "GRAND TOTAL" if co == 1 else (
              f"=SUM({cl}{T1_DATA}:{cl}{T1_DATA + len(clients) - 1})")
        apply_cell(ws, T1_TOT, co, val, bold=True, color="FFFFFF",
                   bg="1F4E79", align_h="center")
        if co > 1:
            ws.cell(T1_TOT, co).number_format = "0"
    return T1_TOT


def _table2_rev_vs_execution(ws, records, last_raw, rev_thr, exe_thr,
                              start_row) -> Tuple[int, int]:
    section_title(ws, start_row, 9,
        "  TABLE 2 :  REVENUE CONTRIBUTION vs EXECUTION CATEGORIZATION")
    T2_HDR = start_row + 1
    header_row(ws, T2_HDR,
               ["Sr.", "Client Name", "Project Reference",
                "Total Work Value (₹)", "Billed Amount (₹)",
                "Revenue Contribution %", "Execution %", "Category"],
               ["1F4E79"] * 8)
    T2_DATA = T2_HDR + 1

    for idx, rec in enumerate(records, 1):
        r      = T2_DATA + idx - 1
        rd_row = idx + 2
        bg     = ALT_ROW_BG if idx % 2 == 0 else "FFFFFF"
        apply_cell(ws, r, 1, idx, bg=bg, align_h="center")
        for col, src in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
            c = ws.cell(r, col, f"='Raw Data'!{src}{rd_row}")
            c.font = _font(); c.border = _THIN_BORDER; c.fill = _fill(bg)
            c.alignment = _align(h="left" if col <= 3 else "right")
            if col in (4, 5):
                c.number_format = "#,##0.00"
        rv = ws.cell(r, 6,
             f"='Raw Data'!D{rd_row}/SUM('Raw Data'!$D$3:$D${last_raw})")
        rv.number_format = "0.00%"; rv.font = _font()
        rv.alignment = _align(h="center"); rv.border = _THIN_BORDER; rv.fill = _fill(bg)
        ex = ws.cell(r, 7, f"='Raw Data'!F{rd_row}")
        ex.number_format = "0.1%"; ex.font = _font()
        ex.alignment = _align(h="center"); ex.border = _THIN_BORDER; ex.fill = _fill(bg)
        cat_c = ws.cell(r, 8, category_excel_formula(
            f"F{r}", f"G{r}", rev_thr, exe_thr))
        cat_c.font = _font(bold=True); cat_c.border = _THIN_BORDER; cat_c.fill = _fill(bg)
        cat_c.alignment = _align(h="center")

    T2_LAST = T2_DATA + len(records) - 1
    T2_TOT  = T2_LAST + 1
    ws.merge_cells(start_row=T2_TOT, start_column=1, end_row=T2_TOT, end_column=3)
    dark_total_row(ws, T2_TOT, 8)
    ws.cell(T2_TOT, 1).value     = "TOTAL"
    ws.cell(T2_TOT, 1).alignment = _align(h="right")
    for col, fml, fmt in [
        (4, f"=SUM(D{T2_DATA}:D{T2_LAST})",     "#,##0.00"),
        (5, f"=SUM(E{T2_DATA}:E{T2_LAST})",     "#,##0.00"),
        (6, f"=SUM(F{T2_DATA}:F{T2_LAST})",     "0.00%"),
        (7, f"=AVERAGE(G{T2_DATA}:G{T2_LAST})", "0.0%"),
    ]:
        ws.cell(T2_TOT, col).value         = fml
        ws.cell(T2_TOT, col).number_format = fmt
        ws.cell(T2_TOT, col).alignment     = _align(h="center")
    ws.cell(T2_TOT, 8).value = "4 Categories"
    return T2_DATA, T2_LAST


def _table3_category_summary(ws, T2_DATA, T2_LAST, start_row):
    section_title(ws, start_row, 9, "  TABLE 3 :  CATEGORY SUMMARY TABLE")
    T3_HDR = start_row + 1
    header_row(ws, T3_HDR,
               ["Category", "No. of Projects", "Total Revenue Share %",
                "Avg. Execution %", "Description"],
               ["1F4E79"] * 5, height=30)
    cat_rng = f"H{T2_DATA}:H{T2_LAST}"
    rev_rng = f"F{T2_DATA}:F{T2_LAST}"
    exe_rng = f"G{T2_DATA}:G{T2_LAST}"
    for ci, cat in enumerate(CATS):
        r          = T3_HDR + 1 + ci
        bg_c, fg_c = CAT_COLORS[cat]
        ws.row_dimensions[r].height = 22
        apply_cell(ws, r, 1, cat,
                   bg=bg_c, color=fg_c, bold=True, align_h="center")
        apply_cell(ws, r, 2, f'=COUNTIF({cat_rng},"{cat}")',
                   bg=bg_c, color=fg_c, bold=True, align_h="center")
        apply_cell(ws, r, 3, f'=SUMIF({cat_rng},"{cat}",{rev_rng})',
                   bg=bg_c, color=fg_c, bold=True, align_h="center", fmt="0.00%")
        apply_cell(ws, r, 4, f'=IFERROR(AVERAGEIF({cat_rng},"{cat}",{exe_rng}),0)',
                   bg=bg_c, color=fg_c, bold=True, align_h="center", fmt="0.0%")
        apply_cell(ws, r, 5, CAT_DESC[cat],
                   bg=bg_c, color=fg_c, size=9, align_h="left", wrap=True)


def _table4_duplicate_clients(ws: Worksheet,
                               pairs: List[DuplicatePair],
                               threshold: int,
                               start_row: int) -> int:
    """
    NEW SECTION — writes the fuzzy duplicate detection results.

    If no duplicates found: writes a single green "No duplicates detected" row.
    If duplicates found:    writes one row per flagged pair with similarity score,
                            a recommendation, and colour-coded severity.

    Returns the next available row after this section.
    """
    SPAN = 9

    # ── Section banner
    section_title(ws, start_row, SPAN,
                  f"  ⚠  TABLE 4 :  POTENTIAL DUPLICATE CLIENT NAMES  "
                  f"(fuzzy similarity ≥ {threshold}%)",
                  bg=DUPE_HDR, fg="FFFFFF")
    ws.cell(start_row, 1).font = Font(
        name="Arial", bold=True, size=11, color="FFFFFF")
    r = start_row + 1

    # ── Explanation note
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
    note = ws.cell(r, 1,
        "These client names scored above the similarity threshold using fuzzy "
        "token matching. Review and merge in the source data if they represent "
        "the same entity. Similarity is computed using rapidfuzz.token_sort_ratio.")
    note.font      = Font(name="Arial", italic=True, size=9, color="595959")
    note.fill      = _fill("FFF9E6")
    note.alignment = _align(h="left", wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1

    if not pairs:
        # ── Clean bill of health
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
        ok = ws.cell(r, 1,
            f"✓  No potential duplicates found among {0} unique client name pairs "
            f"at threshold {threshold}%.")
        ok.font      = _font(bold=True, color="375623", size=11)
        ok.fill      = _fill("C6EFCE")
        ok.alignment = _align(h="center")
        ok.border    = _THIN_BORDER
        ws.row_dimensions[r].height = 24
        return r + 3

    # ── Column headers
    HDR_LABELS = ["#", "Client Name A", "Client Name B",
                  "Similarity %", "Severity", "Recommendation"]
    HDR_BGS    = [DUPE_HDR] * 6
    header_row(ws, r, HDR_LABELS, HDR_BGS, height=30)
    r += 1

    for num, pair in enumerate(pairs, 1):
        bg = DUPE_BG if num % 2 == 0 else "FFFFFF"

        # Severity bands
        if pair.similarity >= 97:
            severity = "LIKELY SAME"
            sev_bg   = "FF0000"
            sev_fg   = "FFFFFF"
            rec      = "Almost certainly the same client — merge immediately."
        elif pair.similarity >= 92:
            severity = "HIGH"
            sev_bg   = "ED7D31"
            sev_fg   = "FFFFFF"
            rec      = "Very likely the same client — verify and consolidate."
        else:
            severity = "MODERATE"
            sev_bg   = "FFEB9C"
            sev_fg   = "9C5700"
            rec      = "Possible duplicate — manual review recommended."

        apply_cell(ws, r, 1, num,           bg=bg, align_h="center", bold=True)
        apply_cell(ws, r, 2, pair.name_a,   bg=bg)
        apply_cell(ws, r, 3, pair.name_b,   bg=bg)
        apply_cell(ws, r, 4, pair.similarity / 100,
                   bg=bg, align_h="center", fmt="0.0%",
                   bold=True, color=WARN_RED)
        apply_cell(ws, r, 5, severity,
                   bg=sev_bg, color=sev_fg, bold=True, align_h="center")
        apply_cell(ws, r, 6, rec,           bg=bg, wrap=True, size=9)
        ws.row_dimensions[r].height = 20
        r += 1

    # ── Summary footer
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
    summary = ws.cell(r, 1,
        f"⚠  {len(pairs)} potential duplicate pair(s) detected.  "
        f"Add entries to Config.client_name_map to normalise names before "
        f"re-running, or increase fuzzy_threshold to suppress false positives.")
    summary.font      = _font(bold=True, size=10, color="9C5700")
    summary.fill      = _fill("FFF2CC")
    summary.alignment = _align(h="left", wrap=True)
    summary.border    = _THIN_BORDER
    ws.row_dimensions[r].height = 32

    return r + 3


def build_analysis_sheet(ws: Worksheet,
                          records: List[ProjectRecord],
                          clients: List[str],
                          last_raw: int,
                          dup_pairs: List[DuplicatePair],
                          cfg: Config) -> Tuple[int, int]:
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2E75B6"

    title_banner(ws, 1, 9, "ANALYSIS DASHBOARD  —  Project Portfolio Intelligence")

    rev_thr, exe_thr = _threshold_params(ws, last_raw)

    t1_tot = _table1_status_distribution(ws, clients, last_raw, start_row=8)

    t2_data, t2_last = _table2_rev_vs_execution(
        ws, records, last_raw, rev_thr, exe_thr, start_row=t1_tot + 3)

    t3_last_row = t2_last + 4 + len(CATS) + 2
    _table3_category_summary(ws, t2_data, t2_last, start_row=t2_last + 4)

    # ── NEW: Table 4 — Duplicate Detection
    _table4_duplicate_clients(ws, dup_pairs, cfg.fuzzy_threshold,
                               start_row=t3_last_row + 3)

    for col, width in [(1,8),(2,38),(3,45),(4,22),(5,22),(6,22),(7,16),(8,22),(9,30)]:
        col_w(ws, col, width)

    return t2_data, t2_last


# ══════════════════════════════════════════════════════════════════════════════
#  ⓫  SHEET 3  —  INSIGHTS  (unchanged logic)
# ══════════════════════════════════════════════════════════════════════════════

def _insights_category_block(ws, cat, row_indices, t2_data,
                              an_cat_range, an_rev_range, an_exe_range,
                              cursor) -> int:
    bg_c, fg_c = CAT_COLORS[cat]

    section_title(ws, cursor, 7, f"  {cat.upper()}", bg=bg_c)
    ws.cell(cursor, 1).font = Font(name="Arial", bold=True, size=12, color=fg_c)
    cursor += 1

    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=7)
    lc = ws.cell(cursor, 1, CAT_DESC[cat])
    lc.font = Font(name="Arial", italic=True, size=9, color="595959")
    lc.fill = _fill("F2F2F2"); lc.alignment = _align(h="left")
    cursor += 1

    header_row(ws, cursor,
               ["Sr.", "Client Name", "Project Reference",
                "Total Work Value (₹)", "Revenue Contribution %", "Execution %"],
               [bg_c] * 6, height=30)
    for col in range(1, 7):
        ws.cell(cursor, col).font = _font(bold=True, color=fg_c)
    cursor += 1

    for slot, rec_idx in enumerate(row_indices, 1):
        an_row = t2_data + (rec_idx - 1)
        bg     = ALT_ROW_BG if slot % 2 == 0 else "FFFFFF"
        apply_cell(ws, cursor, 1, slot, bg=bg, align_h="center", bold=True)
        apply_cell(ws, cursor, 2, f"=Analysis!$B${an_row}", bg=bg, align_h="left")
        apply_cell(ws, cursor, 3, f"=Analysis!$C${an_row}", bg=bg, align_h="left")
        for col, src_col, fmt, h_align in [
            (4, "D", "#,##0.00", "right"),
            (5, "F", "0.00%",   "center"),
            (6, "G", "0.0%",    "center"),
        ]:
            c = ws.cell(cursor, col, f"=Analysis!${src_col}${an_row}")
            c.font = _font(); c.border = _THIN_BORDER
            c.alignment = _align(h=h_align); c.fill = _fill(bg)
            c.number_format = fmt
        cursor += 1

    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=2)
    for co in range(1, 8):
        apply_cell(ws, cursor, co, bg=bg_c, color=fg_c, bold=True, align_h="right")
    ws.cell(cursor, 1).value     = "Projects in category:"
    ws.cell(cursor, 1).alignment = _align(h="right")
    ws.cell(cursor, 3).value         = f'=COUNTIF({an_cat_range},"{cat}")'
    ws.cell(cursor, 3).alignment     = _align(h="center")
    ws.cell(cursor, 4).value         = "Total Rev. Share:"
    ws.cell(cursor, 4).alignment     = _align(h="right")
    ws.cell(cursor, 5).value         = f'=SUMIF({an_cat_range},"{cat}",{an_rev_range})'
    ws.cell(cursor, 5).number_format = "0.00%"
    ws.cell(cursor, 5).alignment     = _align(h="center")
    ws.cell(cursor, 6).value         = "Avg. Execution:"
    ws.cell(cursor, 6).alignment     = _align(h="right")
    ws.cell(cursor, 7).value = (
        f'=IFERROR(AVERAGEIF({an_cat_range},"{cat}",{an_exe_range}),0)')
    ws.cell(cursor, 7).number_format = "0.0%"
    ws.cell(cursor, 7).alignment     = _align(h="center")
    return cursor + 3


def build_insights_sheet(ws: Worksheet,
                          cat_buckets: Dict[str, List[int]],
                          t2_data: int, t2_last: int) -> None:
    ws.sheet_properties.tabColor = "375623"
    title_banner(ws, 1, 7, "INSIGHTS  —  Projects Grouped by Strategic Category")

    an_cat_range = f"Analysis!$H${t2_data}:$H${t2_last}"
    an_rev_range = f"Analysis!$F${t2_data}:$F${t2_last}"
    an_exe_range = f"Analysis!$G${t2_data}:$G${t2_last}"

    cursor = 3
    for cat in CATS:
        cursor = _insights_category_block(
            ws, cat, cat_buckets[cat], t2_data,
            an_cat_range, an_rev_range, an_exe_range, cursor)

    for col, width in [(1,6),(2,38),(3,48),(4,22),(5,22),(6,16),(7,16)]:
        col_w(ws, col, width)


# ══════════════════════════════════════════════════════════════════════════════
#  ⓬  PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def generate_dashboard(cfg: Config = DEFAULT_CONFIG) -> Path:
    """
    Full pipeline:
      load → auto-detect columns → validate → fuzzy-detect duplicates
      → build workbook (3 sheets) → save.

    Parameters
    ----------
    cfg : Config  — all settings; instantiate a custom Config() to call
                    this programmatically without CLI arguments.

    Returns
    -------
    Path : absolute path to the saved output file.
    """
    # 1. Load & auto-detect columns
    records, col_map = load_records(cfg.input_file, cfg)
    log.info("Columns  : client=%d  po=%d  wv=%d  billed=%d",
             col_map.client, col_map.po, col_map.wv, col_map.billed)

    clients     = unique_clients(records)
    cat_buckets = compute_categories(records)

    log.info("Records  : %d  |  Unique clients: %d", len(records), len(clients))
    for cat, rows in cat_buckets.items():
        log.info("  %-24s : %d projects", cat, len(rows))

    # 2. Fuzzy duplicate detection
    dup_pairs = detect_duplicate_clients(clients, cfg.fuzzy_threshold)

    # 3. Build workbook
    wb = Workbook()

    last_raw         = build_raw_data_sheet(wb.active, records)
    t2_data, t2_last = build_analysis_sheet(
        wb.create_sheet("Analysis"),
        records, clients, last_raw, dup_pairs, cfg)
    build_insights_sheet(wb.create_sheet("Insights"),
                         cat_buckets, t2_data, t2_last)

    # 4. Save
    out = Path(cfg.output_file)
    wb.save(out)
    log.info("Saved    : %s", out.resolve())
    return out.resolve()


# ══════════════════════════════════════════════════════════════════════════════
#  ⓭  CLI
# ══════════════════════════════════════════════════════════════════════════════

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Generate 3-sheet Excel portfolio dashboard with fuzzy duplicate detection.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python generate_portfolio_dashboard.py\n"
            "  python generate_portfolio_dashboard.py data.xlsx report.xlsx\n"
            "  python generate_portfolio_dashboard.py data.xlsx --threshold 90\n"
            "  python generate_portfolio_dashboard.py data.xlsx --no-fuzzy\n"
        ),
    )
    p.add_argument("input",       nargs="?", default=None,
                   help="Source Excel file  (default: Raw_data.xlsx)")
    p.add_argument("output",      nargs="?", default=None,
                   help="Output Excel file  (default: Project_Portfolio_Dashboard.xlsx)")
    p.add_argument("--threshold", type=int,  default=None, metavar="N",
                   help="Fuzzy similarity threshold 0–100  (default: 85)")
    p.add_argument("--no-fuzzy",  action="store_true",
                   help="Disable fuzzy duplicate detection entirely")
    p.add_argument("--scan-rows", type=int,  default=None, metavar="N",
                   help="Max rows to scan for header  (default: 10)")
    return p


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT (FIXED)
# ══════════════════════════════════════════════════════════════════════════════

def main(input_file=None, output_file=None):
    cfg = Config()

    # Flask inputs
    if input_file:
        cfg.input_file = input_file
    if output_file:
        cfg.output_file = output_file

    try:
        out = generate_dashboard(cfg)
        print(f"\n✅ Dashboard saved → {out}\n")
    except Exception as exc:
        print(f"Error: {exc}")
        raise


if __name__ == "__main__":
    main()